from django.shortcuts import render, redirect
from django.core.files.storage import FileSystemStorage
from django.conf import settings
import cv2
import re
from openpyxl import Workbook, load_workbook
import os
import easyocr
import nltk
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords

# Download NLTK resources
nltk.download('punkt')
nltk.download('stopwords')

def upload(request):
    if request.method == 'POST' and request.FILES['file']:
        # Delete previous uploaded file if exists
        if 'uploaded_file_url' in request.session:
            previous_file_path = os.path.join(settings.MEDIA_ROOT, request.session['uploaded_file_url'])
            if os.path.exists(previous_file_path):
                os.remove(previous_file_path)

        file = request.FILES['file']
        fs = FileSystemStorage()
        filename = fs.save(file.name, file)
        uploaded_file_url = fs.url(filename)

        # Store the uploaded file URL in session to keep track of it
        request.session['uploaded_file_url'] = filename

        # Process the image and extract data
        extracted_data = process_image(fs.path(filename))

        return render(request, 'bill_detail.html', {
            'uploaded_file_url': uploaded_file_url,
            'address_match': extracted_data.get('address', 'No match'),
            'rr_number': extracted_data.get('rr_number', 'No match'),
            'reading_date': extracted_data.get('reading_date', 'No match'),
            'account_id': extracted_data.get('account_id', 'No match'),
            'consumption': extracted_data.get('consumption', 'No match'),
            'tax': extracted_data.get('tax', 'No match'),
            'net_amount_due': extracted_data.get('net_amount_due', 'No match'),
        })

    return render(request, 'bill_detail.html')

def process_image(image_path):
    image = cv2.imread(image_path)

    # Convert the image to grayscale
    gray_image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    # Apply some preprocessing (optional but can improve OCR accuracy)
    _, processed_image = cv2.threshold(gray_image, 150, 255, cv2.THRESH_BINARY)

    # Initialize EasyOCR reader
    reader = easyocr.Reader(['en'])

    # Perform OCR on the processed image
    result = reader.readtext(processed_image, detail=0)

    # Join the result into a single string
    text = " ".join(result)

    # Print the OCR result to verify
    print("OCR Text:", text)

    # Tokenize text using NLTK
    tokens = word_tokenize(text)

    # Remove stopwords
    stop_words = set(stopwords.words('english'))
    filtered_tokens = [word for word in tokens if word.lower() not in stop_words]

    # Join filtered tokens back into a string
    filtered_text = " ".join(filtered_tokens)

    # Print the filtered OCR text
    print("Filtered OCR Text:", filtered_text)

    # Extract data using regex patterns
    extracted_data = extract_data(filtered_text)

    # Save data to Excel
    save_to_excel(extracted_data)

    return extracted_data

def extract_data(filtered_text):
    # Define regex patterns for specific fields
    address_pattern = re.compile(r'Name Address ([\w\s,()]+)')
    rr_pattern = re.compile(r'RR Number \$?(\d+\.?\d+[A-Za-z]*)')
    acc_pattern = re.compile(r'Account ID \$?(\d+\.?\d*)')
    amt_pattern = re.compile(r'Due\s*:\s*(\d+\s\d+\s[\d.]+)')
    consume_pattern = re.compile(r'Consumption \$?(\d+\.?\d*)')
    tax_pattern = re.compile(r'Tax \$?(\d+\.?\d*)')
    date_pattern = re.compile(r'Reading Date\s*:\s*([\d]{2}\s\/\s[\d]{2}\/\s[\d]{2})')

    # Search for patterns in OCR text
    address = address_pattern.search(filtered_text)
    rr = rr_pattern.search(filtered_text)
    accoun = acc_pattern.search(filtered_text)
    amt = amt_pattern.search(filtered_text)
    consume = consume_pattern.search(filtered_text)
    tax = tax_pattern.search(filtered_text)
    date = date_pattern.search(filtered_text)

    # Prepare extracted data dictionary
    extracted_data = {
        'address': address.group(1) if address else "No match",
        'rr_number': rr.group(1) if rr else "No match",
        'reading_date': date.group(1) if date else "No match",
        'account_id': accoun.group(1) if accoun else "No match",
        'consumption': consume.group(1) if consume else "No match",
        'tax': tax.group(1) if tax else "No match",
        'net_amount_due': amt.group(1) if amt else "No match"
    }

    # Print the matches to verify
    print("Address Match:", extracted_data['address'])
    print("RR Number:", extracted_data['rr_number'])
    print("Reading Date:", extracted_data['reading_date'])
    print("Account ID Match:", extracted_data['account_id'])
    print("Consumption Match:", extracted_data['consumption'])
    print("Tax Match:", extracted_data['tax'])
    print("Net Amount Due Match:", extracted_data['net_amount_due'])

    return extracted_data

def save_to_excel(extracted_data):
    excel_path = os.path.join(settings.MEDIA_ROOT, 'bill.xlsx')
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Bill Information'
        # Write headers if creating a new workbook
        ws['A1'] = 'Name'
        ws['B1'] = 'RR Number'
        ws['C1'] = 'Date'
        ws['D1'] = 'Account ID'
        ws['E1'] = 'Consumption'
        ws['F1'] = 'Tax'
        ws['G1'] = 'Net Amount Due'

    # Determine the next row to write data
    next_row = ws.max_row + 1

    # Write data if found
    if extracted_data['address'] != "No match":
        ws[f'A{next_row}'] = extracted_data['address']
    if extracted_data['rr_number'] != "No match":
        ws[f'B{next_row}'] = extracted_data['rr_number']
    if extracted_data['reading_date'] != "No match":
        ws[f'C{next_row}'] = extracted_data['reading_date']
    if extracted_data['account_id'] != "No match":
        ws[f'D{next_row}'] = extracted_data['account_id']
    if extracted_data['consumption'] != "No match":
        ws[f'E{next_row}'] = extracted_data['consumption']
    if extracted_data['tax'] != "No match":
        ws[f'F{next_row}'] = extracted_data['tax']
    if extracted_data['net_amount_due'] != "No match":
        ws[f'G{next_row}'] = extracted_data['net_amount_due']

    # Save the workbook
    wb.save(excel_path)
    print(f"Workbook saved at {excel_path}")
